import datetime
import pythoncom
import win32com.client as win32
import os
import random
import shutil
import openpyxl

from django import http
from django.forms import widgets
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .formulaire import Ajout_agent
from django.contrib import messages
from APP_LAC import formulaire, models
from django.http import JsonResponse
from docxtpl import DocxTemplate

def mesures_securitaires(request):
    try:
        if request.session:
            request.session['utilisateur']
    except Exception as exc:
        return redirect('index_office')
def index_office(request):
    info = {
        'form': formulaire.Connexion,
        'titre': 'Bienvenue dans le back office',
        'entete': 'LIGUE ANTI CHOMAGE ASBL',
        'title': 'BACK OFFICE'
    }
    return render(request, "office/index.html", context=info)


#affichage de la vue Ajouter un agent
def ajout_agent(request):
    try:
        if request.session['utilisateur']:
            info = {
                'form': Ajout_agent,
                'titre': 'Ajouter un agent',
                'entete': 'LIGUE ANTI CHOMAGE ASBL',
                'date': datetime.datetime.now(),
                'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            }
            return render(request, "office/ajouter_agent.html", context=info)
    except Exception as exc:
        return redirect('index_office')


#création d'un nouvel utilisateur/agent
def create_user(request):
    try:
        if request.session['utilisateur']:
            if request.method == "POST":
                form = Ajout_agent(request.POST, request.FILES)
                nbr = random.randint(1, 1000)
                nbr2 = random.randint(1, 1000)
                if form.is_valid():
                    scan = models.Utilisateur.objects.filter(nom_utilisateur=request.POST['nom_utilisateur'])
                    if scan:
                        return HttpResponse("Ce nom d'utilisateur existe déjà")
                    else:
                        my_user = models.Utilisateur()
                        my_user.nom = request.POST['nom']
                        my_user.postnom = request.POST['postnom']
                        my_user.prenom = request.POST['prenom']
                        my_user.nom_utilisateur = request.POST['nom_utilisateur']
                        my_user.mot_de_passe = request.POST['mot_de_passe']
                        my_user.typeCompte = 1
                        my_user.email = request.POST['email']
                        my_user.tel1 = request.POST['tel1']
                        my_user.tel2 = request.POST['tel2']
                        repository_name = str(request.POST['nom']) + str(request.POST['postnom']) + str(nbr2) + "_" + \
                                          str(request.POST['prenom']) + str(
                            request.POST['nom_utilisateur']) + "_" + str(nbr)
                        my_user.agentFolderPht = "utilisateurs/{}/images".format(repository_name)
                        my_user.agentFolderDoc = "utilisateurs/{}/documents".format(repository_name)
                        if request.FILES.get('photo'):
                            my_user.photo = request.FILES['photo']
                        if request.FILES.get('cv'):
                            my_user.cv = request.FILES['cv']
                        if request.FILES.get('lettre'):
                            my_user.lettre = request.FILES['lettre']
                        if request.FILES.get('contrat'):
                            my_user.contrat = request.FILES['contrat']
                        if request.FILES.get('carteid'):
                            my_user.carteid = request.FILES['carteid']
                        if request.FILES.get('diplome1'):
                            my_user.diplome1 = request.FILES['diplome1']
                        if request.FILES.get('diplome2'):
                            my_user.diplome2 = request.FILES['diplome2']
                        if request.FILES.get('diplome3'):
                            my_user.diplome3 = request.FILES['diplome3']

                        antenne = models.Antenne.objects.get(id=request.POST['antenne'])
                        my_user.antenne = antenne

                        departement = models.Departement.objects.get(id=request.POST['departement'])
                        my_user.departement = departement

                        fonction = models.Fonctions.objects.get(id=request.POST['fonction'])
                        my_user.fonction = fonction
                        my_user.save()

                        return redirect("admindashboard")
            else:
                form = Ajout_agent()
                # pointer vers le fichier du template index.html
                return render(request, 'ajouter_agent.html', {'form': form})
    except Exception as exc:
        return redirect('index_office')


#vue de modification du profile d'un agent par l'administrateur
def profil_agent(request, pk):
    try:
        if request.session['utilisateur']:
            data = models.Utilisateur.objects.get(id=pk)
            form = formulaire.Profile_agent(instance=data)
            info = {
                'id': pk,
                'user': data,
                'currentSession': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
                'form': form,
                'titre': "PROFILE DE L'AGENT",
                'entete': 'LIGUE ANTI CHOMAGE ASBL',
                'date': datetime.datetime.now()
            }
            if request.method == "POST":
                form = formulaire.Profile_agent(request.POST, request.FILES, instance=data)
                if form.is_valid():
                    form.save()
                    return redirect("admindashboard")

            return render(request, "office/profil_agent.html", context=info)
    except Exception as exc:
        return redirect('index_office')


#Affichage du profile de l'utilisateur courant
def mon_profil_agent(request, pk):
    try:
        if request.session['utilisateur']:
            data = models.Utilisateur.objects.get(id=pk)
            suiviContratAgent = {
                'dureeContrat': '',
                'cahierDeCharge': '',
                'salaire': '',
                'formationbourses': ''
            }
            try:
                suiviContratAgent = models.SuiviContrat.objects.get(utilisateur=data)
            except Exception as exc:
                pass
            info = {
                'user': data,
                'suivicontrat': suiviContratAgent,
                'currentSession': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
                'titre': "PROFILE DE L'AGENT",
                'entete': 'LIGUE ANTI CHOMAGE ASBL',
                'date': datetime.datetime.now()
            }
            return render(request, "office/mon_profil.html", context=info)
    except Exception as exc:
        return HttpResponse(exc)
        #return redirect('index_office')


#mise a jour du profile de l'utilisateur courant
def mod_modif_profil(request, pk):
    try:
        if request.session['utilisateur']:
            data = models.Utilisateur.objects.get(id=pk)
            form = formulaire.Mon_profil_agent(instance=data)
            info = {
                'currentSession': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
                'user': data,
                'form': form,
                'titre': "MODIFIER VOTRE PROFILE",
                'entete': 'LIGUE ANTI CHOMAGE ASBL',
                'date': datetime.datetime.now()
            }
            if request.method == "POST":
                if form.is_valid() or 1 == 1:
                    if data.typeCompte != 0:
                        forme = formulaire.Mon_profil_agent(request.POST, request.FILES, instance=data)
                        forme.save()
                    else:
                        forme = formulaire.Mon_profil_agent(request.POST, request.FILES, instance=data)
                        my_model = forme.save(commit=False)
                        my_model.agentFolderPht = "utilisateurs/0_admin/images"
                        my_model.agentFolderDoc = "utilisateurs/0_admin/documents"
                        my_model.save()
                    return redirect("admindashboard")
            return render(request, "office/modif_profil_agent.html", context=info)
    except Exception as exc:
        return redirect('index_office')


def module_connexion(request):
    if request.method == 'POST':
        form = formulaire.Connexion(request.POST)
        if form.is_valid():
            scan = models.Utilisateur.objects.filter(nom_utilisateur=form.data['nom_utilisateur'])
            if len(scan):
                for user in scan:
                    if user.mot_de_passe == form.data['mot_de_passe']:
                        if user.typeCompte == 0:
                            request.session['utilisateur'] = form.data['nom_utilisateur']
                            messages.success(request,
                                             "Bienvenu {} {} {}".format(user.prenom.capitalize(), user.nom.upper(),
                                                                        user.postnom.upper()))
                            return redirect(admindashboard)
                        else:
                            request.session['utilisateur'] = form.data['nom_utilisateur']
                            messages.success(request, "Bienvenu {} {} {}".format(user.prenom, user.nom, user.postnom))
                            return redirect(agentdashboard)
                    else:
                        return redirect(index_office)
            else:
                # le client ne fait pas entrer les bonnes informations pour
                # la connexion
                messages.error(request, "Nom d'utilisateur ou mot de passe incorrect")
                return redirect(index_office)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect(index_office)


def admindashboard(request):
    try:
        if request.session['utilisateur']:
            currentUser = request.session['utilisateur']
            manager = models.Utilisateur.objects.get(nom_utilisateur=currentUser)
            list_agent = models.Utilisateur.objects.all()
            list_note = models.NoteService.objects.all()
            list_echo = models.EchoBureau.objects.all()
            agenda = models.Agenda.objects.filter(utilisateur=manager)
            info = {
                'titre': 'Tableau de Bord',
                'entete': 'LIGUE ANTI CHOMAGE ASBL/DASHBOARD',
                'currentSession': manager,
                'liste_agents': list_agent,
                'liste_notes': list_note,
                'liste_echo': list_echo,
                'agenda': agenda,
                'date': datetime.datetime.now()
            }
            return render(request, "office/dashboard_admin.html", context=info)
    except Exception as exc:
        return HttpResponse(exc)


def agentdashboard(request):
    try:
        if request.session['utilisateur']:
            currentUser = request.session['utilisateur']
            agent = models.Utilisateur.objects.get(nom_utilisateur=currentUser)
            list_echo = models.EchoBureau.objects.all()
            agenda = models.Agenda.objects.filter(utilisateur=agent)
            suiviContratAgent = {
                'dureeContrat': '',
                'cahierDeCharge': '',
                'salaire': '',
                'formationbourses': ''
            }
            try:
                suiviContratAgent = models.SuiviContrat.objects.get(utilisateur=agent)
            except Exception as exc:
                pass
            info = {
                'titre': 'Tableau de Bord',
                'entete': 'LIGUE ANTI CHOMAGE ASBL/DASHBOARD',
                'currentSession': agent,
                'liste_echo': list_echo,
                'suivicontrat': suiviContratAgent,
                'agenda': agenda,
                'date': datetime.datetime.now()
            }
            return render(request, "office/dashboard.html", context=info)
    except Exception as exc:
        return HttpResponse(exc)
        #return redirect('index_office')



def module_deconnexion(request):
    try:
        del request.session['utilisateur']
    except Exception as exc:
        pass
    finally:
        return redirect(index_office)


def note_service(request):
    try:
        if request.session['utilisateur']:
            currentSession = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            form = formulaire.NoteService
            info = {
                'form': form,
                'titre': 'note de service',
                'user': currentSession,
                'date': datetime.datetime.now()
            }
            if request.method == "POST":
                form = formulaire.NoteService(request.POST)
                if form.is_valid():
                    form.save()
                    return redirect('admindashboard')
            return render(request, "office/note_de_service.html", context=info)
    except Exception as exc:
        return redirect('index_office')


#module pour modifier une note de service
def modif_note_service(request, pk):
    try:
        if request.session['utilisateur']:
            note = models.NoteService.objects.get(id=pk)
            form = formulaire.NoteService(instance=note)
            currentSession = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            info = {
                'titre': 'Note de service',
                'form': form,
                'user': currentSession,
                'date': datetime.datetime.now()
            }
            if request.method == "POST":
                if form.is_valid() or 1==1:
                    myform = formulaire.NoteService(request.POST, instance=note)
                    myform.save()
                    return redirect('admindashboard')
            return render(request, "office/note_de_service.html", info)
    except Exception as exc:
        return redirect('index_office')


def mod_sup_note(request, pk):
    try:
        if request.session['utilisateur']:
            note = models.NoteService.objects.get(id=pk)
            note.delete()
            return redirect('admindashboard')
    except Exception as exc:
        return redirect('index_office')


# Fenetre suplementaire Calvin AWA MUSAFIRI (demande de confirmation de suppression d'un agent)
def demande_de_confirmation(request, pk):
    try:
        if request.session['utilisateur']:
            currentSession = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            agent_a_supprimer = models.Utilisateur.objects.get(id=pk)
            info = {
                'titre': 'Demande de confirmation',
                'currentSession': currentSession,
                'date': datetime.datetime.now(),
                'agent_a_supprimer':agent_a_supprimer,
            }
            return render(request, "office/demande_de_confirmation.html", context=info)
    except Exception as exc:
        return redirect('index_office')
    return render(redirect,'office/demande_de_confirmation.html')


# Fenetre suplementaire Calvin AWA MUSAFIRI (demande de confirmation de suppression d'une note de service)
def confirmation_note_de_service(request, pk):
    try:
        if request.session['utilisateur']:
            currentSession = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            note_a_supprimer = models.NoteService.objects.get(id=pk)
            info = {
                'titre': 'Demande de confirmation',
                'currentSession': currentSession,
                'date': datetime.datetime.now(),
                'note_a_supprimer':note_a_supprimer,
            }
            return render(request, "office/confirmation_note.html", context=info)
    except Exception as exc:
        return redirect('index_office')


def effacer_un_agent(request, pk):
    try:
        if request.session['utilisateur']:
            user = models.Utilisateur.objects.get(id=pk)
            # récupération du chemin du repertoire de l'utilisateur à supprimer
            path = user.agentFolderPht[13:]
            index = path.find('/')
            path = path[:index]
            path = 'media/utilisateurs/{}'.format(path)
            #vérifie si le repertoire existe
            if os.path.exists(path):
                shutil.rmtree(path)
            user.delete()
            return redirect("admindashboard")
    except Exception as exc:
        return redirect('index_office')


def effacer_une_note(request, pk):
    models.NoteService.objects.get(id=pk).delete()
    return redirect('admindashboard')


def ajouter_echo_de_bureau(request):
    try:
        if request.session['utilisateur']:
            form = formulaire.EchoBureau
            user = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            info = {
                'form': form,
                'titre': 'Ajouter un echo de bureau',
                'entete': 'LIGUE ANTI CHOMAGE ASBL',
                'date': datetime.datetime.now(),
                'user': user
            }
            if request.method == "POST":
                form = formulaire.EchoBureau(request.POST, request.FILES)
                form.save()
                if user.typeCompte == 0:
                    return redirect('admindashboard')
                else:
                    return redirect('agentdashboard')
            return render(request, "office/ajout_echo_de_bureau.html", context=info)
    except Exception as exc:
        return redirect('index_office')
# AJAX Calvin AWA MUSAFIRI


#suppression d'une publication dans echo de nos bureaux
def supprimer_echo_de_bureau(request, pk):
    models.EchoBureau.objects.get(id=pk).delete()
    return redirect('admindashboard')


def ajouter(request):
    try:
        og = request.GET.get('objglob', None)
        model = models.tester(nom=og)
        model.save()
        return JsonResponse({"success": True})
    except Exception as exc:
        return JsonResponse({"error": False})

def test(request):
    info = {
        'form': formulaire.Test
    }
    return render(request, 'office/test.html', context=info)


def scaner_les_utilisateur(request):
    try:
        return JsonResponse({"success": True})
    except Exception as exc:
            return JsonResponse({"error": False})


def mes_document_admin(request, pk, doc):
    if doc == "cv":
        file = models.Utilisateur.objects.get(id=pk).cv.url
        if file != "":
            file = '{}/{}'.format(os.getcwd(), file)
            f = open(file, 'rb')
            data = f.read()
            return HttpResponse(data, content_type="application/pdf")
        else:
            return HttpResponse("Aucun cv trouvé")

    elif doc == "lettre":
        file = models.Utilisateur.objects.get(id=pk).lettre.url
        if file != "":
            file = '{}/{}'.format(os.getcwd(), file)
            f = open(file, 'rb')
            data = f.read()
            return HttpResponse(data, content_type="application/pdf")
        else:
            return HttpResponse("Aucune lettre trouvée")

    elif doc == "contrat":
        file = models.Utilisateur.objects.get(id=pk).contrat.url
        if file != "":
            file = '{}/{}'.format(os.getcwd(), file)
            f = open(file, 'rb')
            data = f.read()
            return HttpResponse(data, content_type="application/pdf")
        else:
            return HttpResponse("Aucun contrat trouvé")

    elif doc == "carteid":
        file = models.Utilisateur.objects.get(id=pk).carteid.url
        if file != "":
            file = '{}/{}'.format(os.getcwd(), file)
            f = open(file, 'rb')
            data = f.read()
            return HttpResponse(data, content_type="application/pdf")
        else:
            return HttpResponse("Aucune carte d'identité trouvée")

    elif doc == "diplome1":
        file = models.Utilisateur.objects.get(id=pk).diplome1.url
        if file != "":
            file = '{}/{}'.format(os.getcwd(), file)
            f = open(file, 'rb')
            data = f.read()
            return HttpResponse(data, content_type="application/pdf")
        else:
            return HttpResponse("Pas de diplôme trouvé")

    elif doc == "diplome2":
        file = models.Utilisateur.objects.get(id=pk).diplome2.url
        if file != "":
            file = '{}/{}'.format(os.getcwd(), file)
            f = open(file, 'rb')
            data = f.read()
            return HttpResponse(data, content_type="application/pdf")
        else:
            return HttpResponse("Pas de deuxième diplôme trouvé")

    elif doc == "diplome3":
        file = models.Utilisateur.objects.get(id=pk).diplome3.url
        if file != "":
            file = '{}/{}'.format(os.getcwd(), file)
            f = open(file, 'rb')
            data = f.read()
            return HttpResponse(data, content_type="application/pdf")
        else:
            return HttpResponse("Pas de troisième diplôme trouvé")


def ajouter_agenda(request):
    try:
        if request.session['utilisateur']:
            form = formulaire.Agenda
            user = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            info = {
                'form': form,
                'titre': 'Etablir un agenda',
                'entete': 'LIGUE ANTI CHOMAGE ASBL',
                'date': datetime.datetime.now(),
                'user': user
            }
            if request.method == "POST":
                form = formulaire.Agenda(request.POST)
                if form.is_valid():
                    my_form = form.save(commit=False)
                    my_form.utilisateur = user
                    my_form.save()
                    return redirect('agentdashboard')
            return render(request, 'office/agenda.html', info)
    except Exception as exc:
        return redirect('index_office')


def mod_modifier_agenda(request, pk):
    try:
        if request.session['utilisateur']:
            prog = models.Agenda.objects.get(id=pk)
            form = formulaire.Agenda(instance=prog)
            currentSession = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            info = {
                'titre': 'Agenda',
                'form': form,
                'user': currentSession,
                'date': datetime.datetime.now()
            }
            if request.method == "POST":
                if form.is_valid() or 1 == 1:
                    myform = formulaire.Agenda(request.POST, instance=prog)
                    myform.save()
                    if currentSession.typeCompte != 0:
                        return redirect('agentdashboard')
                    else:
                        return redirect('admindashboard')
            return render(request, "office/agenda.html", info)
    except Exception as exc:
        return redirect('index_office')


def mod_sup_agenda(request, pk):
    try:
        if request.session['utilisateur']:
            currentSession = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            models.Agenda.objects.get(id=pk).delete()
            if currentSession.typeCompte != 0:
                return redirect('agentdashboard')
            else:
                return redirect('admindashboard')
    except Exception as exc:
        return redirect('index_office')


def suivi_de_contrat(request, pk):
    try:
        if request.session['utilisateur']:
            currentSession = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            agent = models.Utilisateur.objects.get(id=pk)
            form = formulaire.SuiviDeContrat
            info = {
                'form': form,
                'titre': 'Suivi de contrat',
                'entete': 'LIGUE ANTI CHOMAGE ASBL',
                'date': datetime.datetime.now(),
                'user': currentSession
            }
            if request.method == "POST":
                form = formulaire.SuiviDeContrat(request.POST)
                if form.is_valid():
                    myform = form.save(commit=False)
                    myform.utilisateur = agent
                    myform.save()
                    return redirect('admindashboard')
        return render(request, 'office/suivi_de_contrat.html', context=info)
    except Exception as exc:
        return redirect('index_office')


def bulletinInfo(request,pk):
    mesures_securitaires(request)
    projet = models.Doc_Projet.objects.get(id=pk)
    form_bulletin = formulaire.Bulletin_Info()
    form_photo = formulaire.Photo_Realisation_Sitrep()

    form_bulletin.initialisation(context=pk)
    bulletin = models.BulletinInfo.objects.filter(projet=projet)
    photos = models.PhotoRealisation_Sitrep.objects.all()
    info = {
        'form_bulletin': form_bulletin,
        'form_photo': form_photo,
        'projet':projet,
        'bulletin':bulletin,
        'photos':photos,
    }
    return render(request,'office/docprojet/bulletin_info/bulletin_information.html',context=info)


def doc_projet(request):
    try:
        if request.session['utilisateur']:
            user = models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
            domaine = models.Domaine.objects.all()
            info = {
                'form': formulaire.Doc_Projet,
                'domaines': domaine,
                'titre': 'Etablir le document du nouveau projet',
                'entete': 'LIGUE ANTI CHOMAGE ASBL',
                'date': datetime.datetime.now(),
                'user': user
            }
            return render(request, "office/document_projet.html", context=info)
    except Exception as exc:
        return redirect('index_office')


def mod_ajouter_doc_projet(request):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Doc_Projet(request.POST)
        if form.is_valid():
            myform = form.save(commit=False)
            myform.code_project = "LAC-" + request.POST['titre'] + "-" + request.POST['finance_par'] + "-" + request.POST['programme_solicite']
            myform.save()
            return redirect('liste_projet')
    else:
        return redirect('index_office')


def liste_projet(request):
    mesures_securitaires(request)
    les_projets = models.Doc_Projet.objects.all()
    info = {
        'projets': les_projets,
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur'])
    }
    return render(request, 'office/docprojet/liste_projet.html', context=info)


def apercu_du_projet(request,pk):
    projet = models.Doc_Projet.objects.get(id=pk)
    mis_a_jour_duree_projet(projet=pk)
    domaines = models.DomaineProjet.objects.filter(projet=projet)
    # detaildomaineprojet = models.DomaineProjetDetail.objects.all()
    volet = models.VoletProjet.objects.all()
    activite_principale = models.ActivitePrincipal.objects.all()
    indicateurs = models.Suivi_de_indicateurs_projet.objects.all()
    gestion_risques = models.Gestion_des_risques_SI.objects.all()
    # sous_activite = models.SousActivite.objects.all()

    #partenaireprojet
    l_partenaire = models.PartenaireProjet.objects.filter(projet=projet)

    #localisation
    pays = models.Pays_Projet.objects.filter(projet=projet)
    provinces = models.Province_Projet.objects.all()
    ville_territoire = models.Ville_Territoire_Projet.objects.all()
    axes = models.Axe_Projet.objects.all()

    #bureau d'execution lac
    cn = models.Coordination_Nationale_Projet.objects.filter(projet=projet)
    cp = models.Coordination_Provinciale_Projet.objects.all()
    aldpe = models.Antenne_Projet.objects.all()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'l_partenaire': l_partenaire,
        'projet':projet,
        'domaines':domaines,
        'volets':volet,
        'activites_principales':activite_principale,
        'pays':pays,
        'provinces':provinces,
        'ville_territoires':ville_territoire,
        'axes':axes,

        'coord_nat':cn,
        'coord_prov':cp,
        'antenne':aldpe,

        'suivi_indicateur':indicateurs,
        'gestion_risques':gestion_risques,

        # 'sousactivites':sous_activite,
        # 'detaildomaineprojet': detaildomaineprojet,
    }
    return render(request, 'office/docprojet/apercu.html', context=info)

#EDIT DOC PROJET
def edit_doc_projet(request,pk):
    mesures_securitaires(request)
    projet = models.Doc_Projet.objects.get(id=pk)
    form = formulaire.Doc_Projet(instance=projet)
    info = {
        'form':form,
        'projet':projet,
    }
    return render(request,'office/docprojet/edit_doc_projet.html',context=info)
def mod_modifier_doc_projet(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        projet = models.Doc_Projet.objects.get(id=pk)
        form = formulaire.Doc_Projet(instance=models.Doc_Projet.objects.get(id=pk),
                                     data=request.POST)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")


def ajouter_volet_domaine(request, pk):
    mesures_securitaires(request)
    domaine = models.DomaineProjet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=domaine.projet.id)
    form = formulaire.VoletProjet(key=pk)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form': form,
        'projet': projet,
    }
    return render(request, 'office/docprojet/ajouter_volet_domaine.html', context=info)


def mod_ajouter_volet_domaine(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        projet = models.Doc_Projet.objects.get(id=pk)
        scan = models.VoletProjet.objects.filter(domaine=request.POST['domaine'], volet=request.POST['volet'])
        if scan:
            return HttpResponse('Ce volet existe déjà pour ce projet')
        else:
            form = formulaire.VoletProjet(request.POST)
            if form.is_valid():
                form.save()
                return redirect('apercu_du_projet', projet.id)
            else:
                return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")


def ajouter_ap_volet(request, pk):
    mesures_securitaires(request)
    volet = models.VoletProjet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=volet.domaine.projet.id)
    form = formulaire.ActivitePrincipale(key=pk)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form': form,
        'projet': projet,
        'volet': volet,
    }
    return render(request, 'office/docprojet/ajouter_activite_principale_volet.html', context=info)


def mod_ajouter_ap_volet(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        volet = models.VoletProjet.objects.get(id=pk)
        projet = models.Doc_Projet.objects.get(id=volet.domaine.projet.id)
        scan = models.ActivitePrincipal.objects.filter(volet=request.POST['volet'], activitePrincipale=request.POST['activitePrincipale'])
        if scan:
            return HttpResponse('Une activité principale avec la même désignation a été trouvé')
        else:
            form = formulaire.ActivitePrincipale(request.POST)
            if form.is_valid():
                form.save()
                return redirect('apercu_du_projet', projet.id)
            else:
                return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")


def mod_supprimer_domaine_resultat(request, pk):
    mesures_securitaires(request)
    domaine = models.DomaineProjet.objects.get(id=pk)
    domaine.delete()
    return redirect("apercu_du_projet", domaine.projet.id)


def mod_effacer_volet(request, pk):
    mesures_securitaires(request)
    volet = models.VoletProjet.objects.get(id=pk)
    volet.delete()
    return redirect("apercu_du_projet", volet.domaine.projet.id)



#EDIT PAYS PROJET
def edit_pays_projet(request,pk):
    mesures_securitaires(request)
    pays = models.Pays_Projet.objects.get(id=pk)
    # projet = models.Doc_Projet.objects.get(id=pk)
    form = formulaire.Pays_Projet(instance=pays)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form':form,
        'pays':pays,
    }
    return render(request,'office/docprojet/edit_pays_projet.html', context=info)

def mod_modifier_pays_projet(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Pays_Projet(instance=models.Pays_Projet.objects.get(id=pk),
                                     data=request.POST)
        pays = models.Pays_Projet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            pays.projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")

def mod_delete_pays_projet(request,pk):
    mesures_securitaires(request)
    pays = models.Pays_Projet.objects.get(id=pk)
    pays.delete()
    return redirect("apercu_du_projet", pays.projet.id)

def mod_delete_province_projet(request,pk):
    mesures_securitaires(request)
    province = models.Province_Projet.objects.get(id=pk)
    province.delete()
    return redirect("apercu_du_projet", province.pays.projet.id)

def mod_delete_villeterritoire_projet(request,pk):
    mesures_securitaires(request)
    villeterritoire = models.Ville_Territoire_Projet.objects.get(id=pk)
    villeterritoire.delete()
    return redirect("apercu_du_projet", villeterritoire.province.pays.projet.id)

def mod_delete_axe_projet(request,pk):
    mesures_securitaires(request)
    axe = models.Axe_Projet.objects.get(id=pk)
    axe.delete()
    return redirect("apercu_du_projet", axe.ville_territoire.province.pays.projet.id)

def mod_delete_coonal_projet(request,pk):
    mesures_securitaires(request)
    coonal = models.Coordination_Nationale_Projet.objects.get(id=pk)
    coonal.delete()
    return redirect("apercu_du_projet", coonal.projet.id)

def mod_delete_coopr_projet(request,pk):
    mesures_securitaires(request)
    coopr = models.Coordination_Provinciale_Projet.objects.get(id=pk)
    coopr.delete()
    return redirect("apercu_du_projet", coopr.cordination_nationale.projet.id)

def mod_delete_coopr_projet(request,pk):
    mesures_securitaires(request)
    coopr = models.Coordination_Provinciale_Projet.objects.get(id=pk)
    coopr.delete()
    return redirect("apercu_du_projet", coopr.cordination_nationale.projet.id)

def mod_delete_antenne_projet(request,pk):
    mesures_securitaires(request)
    antenne = models.Antenne_Projet.objects.get(id=pk)
    antenne.delete()
    return redirect("apercu_du_projet", antenne.coordination_provinciale.cordination_nationale.projet.id)


def edit_partenaire_execution(request, pk):
    mesures_securitaires(request)
    partenaire = models.PartenaireProjet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=partenaire.projet.id)
    form = formulaire.PartenaireProjet(instance=partenaire)

    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form': form,
        'partenaire': partenaire,
        'projet': projet,
        'nom_part': partenaire.partenaire,
    }
    return render(request, 'office/docprojet/edit_partenaire_execution.html', context=info)

def mod_edit_partenaire_execution(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.PartenaireProjet(instance=models.PartenaireProjet.objects.get(id=pk),
                                     data=request.POST)
        partenaire = models.PartenaireProjet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            bg = models.BudgetEtPlanDecaissement.objects.all()
            for element in bg:
                if element.activite_principale.volet.domaine.projet == partenaire.projet:
                    for part in element.partenaire:
                        if part[0] == request.POST['name_part']:
                            part[0] = request.POST['partenaire']
                            element.save()
            return redirect('apercu_du_projet',
                            partenaire.projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")

def delete_partenaire_execution(request,pk):
    mesures_securitaires(request)
    partenaire = models.PartenaireProjet.objects.get(id=pk)
    bg = models.BudgetEtPlanDecaissement.objects.all()
    ma_nouvelle_liste = []
    for element in bg:
        if element.activite_principale.volet.domaine.projet == partenaire.projet:
            for part in element.partenaire:
                if part[0] != partenaire.partenaire:
                     ma_nouvelle_liste.append(part)
            element.partenaire = ma_nouvelle_liste
            element.save()
            ma_nouvelle_liste.clear()
    partenaire.delete()
    return redirect("apercu_du_projet", partenaire.projet.id)

#EDIT PROVINCE PROJET
def edit_province_projet(request,pk,projet):
    mesures_securitaires(request)
    province = models.Province_Projet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=projet)
    form = formulaire.Province_Projet(instance=province)
    form.fields['pays'].widget = widgets.HiddenInput()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form':form,
        'province':province,
        'projet':projet,
    }
    return render(request,'office/docprojet/edit_province_projet.html', context=info)

def mod_modifier_province_projet(request,pk,projet):
    mesures_securitaires(request)
    if request.method == 'POST':
        projet = models.Doc_Projet.objects.get(id=projet)
        form = formulaire.Province_Projet(instance=models.Province_Projet.objects.get(id=pk),
                                     data=request.POST)
        province = models.Province_Projet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")



#EDIT VILLE TERRITOIR PROJET
def edit_territoir_ville_projet(request,pk,projet):
    mesures_securitaires(request)
    ville_territoir = models.Ville_Territoire_Projet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=projet)
    form = formulaire.VilleTerritoir_Projet(instance=ville_territoir)
    form.fields['province'].widget = widgets.HiddenInput()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form':form,
        'ville_territoir':ville_territoir,
        'projet':projet,
    }
    return render(request,'office/docprojet/edit_territoir_ville_projet.html',context=info)
def mod_modifier_territoir_ville_projet(request,pk,projet):
    mesures_securitaires(request)
    if request.method == 'POST':
        projet = models.Doc_Projet.objects.get(id=projet)
        form = formulaire.VilleTerritoir_Projet(instance=models.Ville_Territoire_Projet.objects.get(id=pk),
                                     data=request.POST)
        # ville_territoir = models.Ville_Territoire_Projet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")


#EDIT AXE PROJET
def edit_axe_projet(request,pk,projet):
    mesures_securitaires(request)
    axe = models.Axe_Projet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=projet)
    form = formulaire.Axe_Projet(instance=axe)
    form.fields['ville_territoire'].widget = widgets.HiddenInput()
    info = {
        'form':form,
        'axe':axe,
        'projet':projet,
    }
    return render(request, 'office/docprojet/edit_axe_projet.html', context=info)

def mod_modifier_axe_projet(request,pk,projet):
    mesures_securitaires(request)
    if request.method == 'POST':
        projet = models.Doc_Projet.objects.get(id=projet)
        form = formulaire.Axe_Projet(instance=models.Axe_Projet.objects.get(id=pk),
                                     data=request.POST)
        # ville_territoir = models.Ville_Territoire_Projet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")



#EDIT AXE PROJET
def edit_axe_projet(request,pk,projet):
    mesures_securitaires(request)
    axe = models.Axe_Projet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=projet)
    form = formulaire.Axe_Projet(instance=axe)
    form.fields['ville_territoire'].widget = widgets.HiddenInput()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form':form,
        'axe':axe,
        'projet':projet,
    }
    return render(request,'office/docprojet/edit_axe_projet.html',context=info)
def mod_modifier_axe_projet(request,pk,projet):
    mesures_securitaires(request)
    if request.method == 'POST':
        projet = models.Doc_Projet.objects.get(id=projet)
        form = formulaire.Axe_Projet(instance=models.Axe_Projet.objects.get(id=pk),
                                     data=request.POST)
        # ville_territoir = models.Ville_Territoire_Projet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")

#EDIT COORDINATION NATIONALE PROJET
def edit_cn_projet(request,pk):
    mesures_securitaires(request)
    cn = models.Coordination_Nationale_Projet.objects.get(id=pk)
    form = formulaire.Coordination_National(instance=cn)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form':form,
        'cn':cn,
    }
    return render(request,'office/docprojet/edit_cn_projet.html',context=info)
def mod_modifier_cn_projet(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        cn = models.Coordination_Nationale_Projet.objects.get(id=pk)
        form = formulaire.Coordination_National(instance=models.Coordination_Nationale_Projet.objects.get(id=pk),
                                     data=request.POST)
        # ville_territoir = models.Ville_Territoire_Projet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            cn.projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")

#EDIT COORDINATION PROVINCIALE PROJET
def edit_cp_projet(request,pk,projet):
    mesures_securitaires(request)
    cp = models.Coordination_Provinciale_Projet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=projet)
    form = formulaire.Coordination_Provinciale(instance=cp)
    form.fields['cordination_nationale'].widget = widgets.HiddenInput()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form':form,
        'cp':cp,
        'projet':projet,
    }
    return render(request,'office/docprojet/edit_cp_projet.html',context=info)

def mod_modifier_cp_projet(request,pk,projet):
    mesures_securitaires(request)
    if request.method == 'POST':
        projet = models.Doc_Projet.objects.get(id=projet)
        form = formulaire.Coordination_Provinciale(instance=models.Coordination_Provinciale_Projet.objects.get(id=pk),
                                     data=request.POST)
        # ville_territoir = models.Ville_Territoire_Projet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            projet.id)
        else:
            return HttpResponse("Erreur de formulaire")
    else:
        return redirect("index_office")


#EDIT ANTENNE PROJET
def edit_antenne_projet(request, pk, projet):
    mesures_securitaires(request)
    ant = models.Antenne_Projet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=projet)
    form = formulaire.Antenne_projet(instance=ant)
    form.fields['coordination_provinciale'].widget = widgets.HiddenInput()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form':form,
        'antenne':ant,
        'projet':projet,
    }
    return render(request, 'office/docprojet/edit_antenne_projet.html', context=info)
def mod_modifier_antenne_projet(request,pk,projet):
    mesures_securitaires(request)
    if request.method == 'POST':
        projet = models.Doc_Projet.objects.get(id=projet)
        form = formulaire.Antenne_projet(instance=models.Antenne_Projet.objects.get(id=pk),
                                     data=request.POST)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            projet.id)
        else:
            return HttpResponse("Erreur de formulaire {}".format(form.errors))
    else:
        return redirect("index_office")


# DOMAINE DE RESULTAT
#VOLET
def edit_volet_projet(request,pk):
    mesures_securitaires(request)
    volet = models.VoletProjet.objects.get(id=pk)
    domaine = models.DomaineProjet.objects.get(id=volet.domaine.id)
    form = formulaire.VoletProjet(instance=volet, key=domaine.id)
    form.fields['domaine'].widget = widgets.HiddenInput()
    form.fields['volet'].widget = widgets.HiddenInput()
    info = {
        'projet': domaine.projet,
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form': form,
        'volet': volet,
    }
    return render(request, 'office/docprojet/modifier_volet.html', context=info)

def mod_modifier_volet_projet(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        volet = models.VoletProjet.objects.get(id=pk)
        form = formulaire.VoletProjet(instance=volet, data=request.POST)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            volet.domaine.projet.id)
        else:
            return HttpResponse("Erreur de formulaire {}".format(form.errors))
    else:
        return redirect("index_office")


# ACTUVITE PRINCIPALES
def edit_activite_principale_projet(request, pk):
    mesures_securitaires(request)
    ap = models.ActivitePrincipal.objects.get(id=pk)
    form = formulaire.ActivitePrincipale(instance=ap, key=ap.volet.id)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form': form,
        'projet': ap.volet.domaine.projet,
        'activite_principale': ap,
    }
    return render(request, 'office/docprojet/modifier_ap.html', context=info)

def mod_modifier_activite_principale_projet(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        ap = models.ActivitePrincipal.objects.get(id=pk)
        form = formulaire.ActivitePrincipale(instance=ap, data=request.POST)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet', ap.volet.domaine.projet.id)
        else:
            return HttpResponse("Erreur de formulaire {}".format(form.errors))
    else:
        return redirect("index_office")

# INDICATEURS
def edit_indicateur_projet(request,pk):
    mesures_securitaires(request)
    ind = models.Suivi_de_indicateurs_projet.objects.get(id=pk)
    form = formulaire.Suivi_des_indicateurs(instance=ind, initial={
        'periode_execution':ind.periode_execution
    })
    # form.fields['domaine'].widget = widgets.HiddenInput()
    # form.fields['volet'].widget = widgets.HiddenInput()
    info = {
        'form':form,
        'indicateur':ind,
    }
    return render(request,'office/docprojet/edit_indicateur.html',context=info)

def mod_modifier_indicateur_projet(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        suivi = models.Suivi_de_indicateurs_projet.objects.get(id=pk)
        form = formulaire.Suivi_des_indicateurs(request.POST, instance=suivi)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet',
                            suivi.activite_principale.volet.domaine.projet.id)
        else:
            return HttpResponse("Erreur de formulaire {}".format(form.errors))
    else:
        return redirect("index_office")

# GESTION DES RISQUES
def mod_modifier_risque(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        risque = models.Gestion_des_risques_SI.objects.get(id=pk)
        form = formulaire.Gestion_des_risques(request.POST, instance=risque)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet', risque.volet.domaine.projet.id)
        else:
            return HttpResponse("Erreur de formulaire {}".format(form.errors))
    else:
        return redirect("index_office")


#AP
def modifier_ap(request,pk,idprojet):
    mesures_securitaires(request)
    activite_principale = models.ActivitePrincipal.objects.get(id=pk)
    form = formulaire.ActivitePrincipale(instance=activite_principale)
    info = {
        'form': form,
        'activite_principale': activite_principale,
        'idprojet': idprojet
    }
    return render(request,'office/docprojet/modifier_ap.html',context=info)


def mod_modifier_ap(request,pk,idprojet):
    mesures_securitaires(request)
    if request.method == 'POST':
        activte_principale = models.ActivitePrincipal.objects.get(id=pk)
        form = formulaire.ActivitePrincipale(request.POST,instance=activte_principale)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet', idprojet)
        else:
            return HttpResponse(""" Vous ne pouvez pas l'insérer """)
    else:
        return redirect('index_office')


def mod_effacer_ap(request, pk):
    mesures_securitaires(request)
    try:
        ap = models.ActivitePrincipal.objects.get(id=pk)
        ap.delete()
        return redirect('apercu_du_projet', ap.volet.domaine.projet.id)
    except Exception as exc:
        return HttpResponse(exc)


#OS
# def modifier_os(request, pk, idprojet):
#     mesures_securitaires(request)
#     objectif_specifique = models.ObjectifSpec.objects.get(id=pk)
#     form = formulaire.ObjSpecifique(instance=objectif_specifique)
#     info = {
#         'baniere': 'Modifier un objectif specifique',
#         'form': form,
#         'objectif_specifique': objectif_specifique,
#         'idprojet': idprojet
#     }
#     return render(request,'office/docprojet/modifier_os.html',context=info)


def mod_modifier_os(request,pk,idprojet):
    mesures_securitaires(request)
    if request.method == 'POST':
        objectif_specifique = models.ObjectifSpec.objects.get(id=pk)
        form = formulaire.ObjSpecifique(request.POST,instance=objectif_specifique)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet', idprojet)
        else:
            return HttpResponse("""Vous ne pouvez pas l'inserer """)
    else:
        return redirect('index_office')


def mod_effacer_os(request,pk,idprojet):
    mesures_securitaires(request)
    models.ObjectifSpec.objects.get(id = pk).delete()
    return redirect('apercu_du_projet', idprojet)

#OG
def modifier_og(request,pk,idprojet):
    mesures_securitaires(request)
    objectif_global = models.Objectif_LAC.objects.get(id=pk)
    form = formulaire.Objectif_LAC(instance=objectif_global)
    info = {
        'baniere': 'Modifier un objectif global',
        'form': form,
        'objectif_global': objectif_global,
        'idprojet': idprojet
    }
    return render(request,'office/docprojet/modifier_og.html',context=info)



def mod_modifier_og(request,pk,idprojet):
    mesures_securitaires(request)
    if request.method == 'POST':
        objectif_global = models.Objectif_LAC.objects.get(id=pk)
        form = formulaire.Objectif_LAC(request.POST, instance=objectif_global)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet', idprojet)
        else:
            return HttpResponse("""Vous ne pouvez pas l'inserer """)
    else:
        return redirect('index_office')



def mod_effacer_og(request,pk,idprojet):
    mesures_securitaires(request)
    models.Objectif_LAC.objects.get(id = pk).delete()
    return redirect('apercu_du_projet', idprojet)


#DOMAINES
def modifier_domaine(request, pk):
    mesures_securitaires(request)
    domaineprojet = models.DomaineProjet.objects.get(id=pk)
    projet = models.Doc_Projet.objects.get(id=domaineprojet.projet.id)
    form = formulaire.DomaineProjetEdit(instance=domaineprojet)
    info = {
        "domaineprojet": domaineprojet,
        "user": models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        "projet": projet,
        "form": form,
    }
    return render(request, 'office/docprojet/modifier_domaine.html', context=info)


def mod_modifier_domaine(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        domaine = models.DomaineProjet.objects.get(id=pk)
        projet = models.Doc_Projet.objects.get(id=domaine.projet.id)
        form = formulaire.DomaineProjetEdit(request.POST, instance=domaine)
        if form.is_valid():
            # myform = form.save(commit=False)
            # myform.domaine = domaine
            form.save()
            return redirect('apercu_du_projet', projet.id)
        else:
            return HttpResponse(form.errors)
    else:
        return redirect('index_office')



def mod_effacer_domaine(request,pk,idprojet):
    mesures_securitaires(request)
    models.DomaineProjet.objects.get(id = pk).delete()
    return redirect('apercu_du_projet', idprojet)

def configuration_doc_projet(request, pk):
    mesures_securitaires(request)
    domaine = models.Domaine.objects.all()
    projet_en_cour = models.Doc_Projet.objects.get(id=pk)
    l_partenaire = models.PartenaireProjet.objects.filter(projet=projet_en_cour)
    form_partenaire = formulaire.PartenaireProjet(key=projet_en_cour.id)
    form_province = formulaire.Province_Projet()
    form_province.initialisation(context=pk)
    form_territoir = formulaire.VilleTerritoir_Projet()
    form_territoir.initialisation(context=pk)
    form_axe = formulaire.Axe_Projet()
    form_axe.initialisation(context=pk)

    form_cp = formulaire.Coordination_Provinciale()
    form_cp.initialisation(context=pk)

    form_ald = formulaire.Antenne_projet()
    form_ald.initialisation(context=pk)
    pays = models.Pays_Projet.objects.filter(projet=projet_en_cour)
    province = models.Province_Projet.objects.all()
    ville_territoir = models.Ville_Territoire_Projet.objects.all()
    axes = models.Axe_Projet.objects.all()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'partenaire': form_partenaire,
        'l_partenaire': l_partenaire,
        'projet_en_cour':projet_en_cour,
        'domaines':domaine,
        'form_pays': formulaire.Pays_Projet(),
        'form_province': form_province,
        'form_villet_territoir': form_territoir,
        'form_axe': form_axe,
        'form_cp': form_cp,
        'form_ant': form_ald,
        'pays': pays,
        'province': province,
        'ville_territoir': ville_territoir,
        'axe': axes,

        'form_CN': formulaire.Coordination_National(),
        'form_CN': formulaire.Coordination_National(),
        'cn': models.Coordination_Nationale_Projet.objects.filter(projet=projet_en_cour),
        'cp': models.Coordination_Provinciale_Projet.objects.all(),
        'alde': models.Antenne_Projet.objects.all(),
    }
    return render(request, 'office/docprojet/configuration_doc_projet.html', context=info)


def configuration_domaine_doc_projet(request, pk):
    mesures_securitaires(request)
    domaine = models.Domaine.objects.all()
    projet_en_cour = models.Doc_Projet.objects.get(id=pk)
    Domaine_projet = models.DomaineProjet.objects.filter(projet=pk)

    #affichage ergonomique des elements
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'projet_en_cour': projet_en_cour,
        'domaines': domaine,
        'volets': models.Volet.objects.all(),
        'domaine_projet': Domaine_projet,
        'form_domaine_projet': formulaire.DomaineProjet(),

    }
    return render(request,'office/docprojet/domaine_resultat.html',context=info)


def add_partenaire_execution(request, pk):
    mesures_securitaires(request)
    projet = models.Doc_Projet.objects.get(id=pk)
    if request.method == 'POST':
        form = formulaire.PartenaireProjet(request.POST)
        if form.is_valid():
            scan = models.PartenaireProjet.objects.filter(partenaire=request.POST['partenaire'], projet=pk)
            if len(scan) == 0:
               form.save()
               list = []
               budget = models.BudgetEtPlanDecaissement.objects.all()
               for bg in budget:
                   if bg.activite_principale.volet.domaine.projet == projet:
                       list.extend(bg.partenaire)
                       list.append([request.POST['partenaire'], "0"])
                       bg.partenaire = list
                       bg.save()
                       list.clear()
               return redirect('configuration_doc_projet', pk)
            else:
                return HttpResponse("Rédondance")
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")


def loc_pays(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Pays_Projet(request.POST)
        if form.is_valid():
            scan = models.Pays_Projet.objects.filter(pays=request.POST['pays'],projet=pk)
            if len(scan) == 0:
                myForm = form.save(commit=False)
                myForm.projet = models.Doc_Projet.objects.get(id=pk)
                myForm.save()
                return redirect('configuration_doc_projet', pk)
            else:
                return HttpResponse("Redodance")
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")

def loc_province(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Province_Projet(request.POST)
        if form.is_valid():
            form.save()
            return redirect('configuration_doc_projet', pk)
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")

def loc_ville_territoire(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.VilleTerritoir_Projet(request.POST)
        if form.is_valid():
            form.save()
            return redirect('configuration_doc_projet', pk)
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")

def loc_axe(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Axe_Projet(request.POST)
        if form.is_valid():
            form.save()
            return redirect('configuration_doc_projet', pk)
        else:
            return HttpResponse("erreur")
    else:
        return redirect("index_office")


def mod_ajouter_domainProjet(request, idprojet):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.DomaineProjet(request.POST)
        if form.is_valid():
            myform = form.save(commit=False)
            projet = models.Doc_Projet.objects.get(id=idprojet)
            domaine = models.Domaine.objects.get(domaine=myform.domaine)
            test = models.DomaineProjet.objects.filter(projet=projet, domaine=domaine)
            if test:
                return HttpResponse("Ce domaine existe déjà pour ce projet")
            else:
                myform.projet = projet
                myform.save()
                return redirect('configuration_doc_projet', idprojet)
    else:
        return redirect('configuration_doc_projet', idprojet)

def loc_pays(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Pays_Projet(request.POST)
        if form.is_valid():
            scan = models.Pays_Projet.objects.filter(pays=request.POST['pays'],projet=pk)
            if len(scan) == 0:
                myForm = form.save(commit=False)
                myForm.projet = models.Doc_Projet.objects.get(id=pk)
                myForm.save()
                return redirect('configuration_doc_projet', pk)
            else:
                return HttpResponse("Redodance")
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")


def ajouter_cn(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Coordination_National(request.POST)
        if form.is_valid():
            scan = models.Coordination_Nationale_Projet.objects.filter(coordination_national=request.POST['coordination_national'], projet=pk)
            if len(scan) == 0:
                myForm = form.save(commit=False)
                myForm.projet = models.Doc_Projet.objects.get(id=pk)
                myForm.save()
                return redirect('configuration_doc_projet', pk)
            else:
                return HttpResponse("Redodance")
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")
def ajouter_cp(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Coordination_Provinciale(request.POST)
        if form.is_valid():
            form.save()
            return redirect('configuration_doc_projet', pk)
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")
def ajouter_aldpe(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Antenne_projet(request.POST)
        if form.is_valid():
            form.save()
            return redirect('configuration_doc_projet', pk)
        else:
            return HttpResponse("erreur")
    else:
        return redirect("index_office")


# def domaine_resultat(request,pk):
#     projet = models.Doc_Projet.objects.get(id=pk)
#     form_domaine_resultat = formulaire.Domaine_Resulat()
#     form_liste_personnel = formulaire.Liste_personnel_cle()
#     form_liste_personnel.initialisation(context=pk)
#     domaine_resultat_objects = models.Domaine_Resultat.objects.filter(projet=projet)
#     liste_personnel = models.Liste_des_personnes_SI.objects.all()
#     form_suivi_des_indicateurs = formulaire.Suivi_des_indicateurs()
#     info = {
#         'projet':projet,
#         'domaine_resultat':form_domaine_resultat,
#         'domaines':domaine_resultat_objects,
#         'form_liste_personnel':form_liste_personnel,
#         'liste_personnel':liste_personnel,
#
#         'form_indicateurs': form_suivi_des_indicateurs,
#     }
#     return render(request,'office/docprojet/domaine_resultat.html',context=info)


def mod_ajouter_une_activite_principale(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.ActivitePrincipale(request.POST)
        if form.is_valid():
            form.save()
            return redirect('configuration_doc_projet', pk)
    else:
        return redirect('index_office')
    return redirect('configuration_doc_projet', pk)


def suivi_des_indicateurs(request, pk):
    mesures_securitaires(request)
    ap = models.ActivitePrincipal.objects.get(id=pk)
    form_suivi_des_indicateurs = formulaire.Suivi_des_indicateurs(key=pk)

    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'projet': ap.volet.domaine.projet,
        'form': form_suivi_des_indicateurs,
        'ap': ap,
    }
    return render(request, "office/docprojet/suivi_des_indicateurs.html", context=info)


def mod_ajouter_suivi_indicateur(request, pk):
    mesures_securitaires(request)
    ap = models.ActivitePrincipal.objects.get(id=pk)
    if request.method == 'POST':
        form = formulaire.Suivi_des_indicateurs(request.POST)
        if form.is_valid():
            form.save()
            return redirect('suivi_des_indicateurs', ap.id)
        else:
            return HttpResponse('Erreur formulaire {} {}'.format(form.errors, request.POST))
    else:
        return redirect('index_office')


def ajouter_gestion_risque(request, pk):
    mesures_securitaires(request)
    form = formulaire.Gestion_des_risques(key=pk)
    volet = models.VoletProjet.objects.get(id=pk)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'projet': volet.domaine.projet,
        'form': form,
        'volet': volet,
    }
    return render(request, "office/docprojet/ajouter_gestion_risque_volet.html", context=info)


def mod_ajouter_risque(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        volet = models.VoletProjet.objects.get(id=pk)
        form = formulaire.Gestion_des_risques(request.POST)
        if form.is_valid():
            form.save()
            return redirect('apercu_du_projet', volet.domaine.projet.id)
        else:
            return HttpResponse(form.errors)
    else:
        return redirect('index_office')



def edit_indicateur(request, pk):
    try:
        mesures_securitaires(request)
        indicateur = models.Suivi_de_indicateurs_projet.objects.get(id=pk)

        form = formulaire.Suivi_des_indicateurs(instance=indicateur, key=indicateur.activite_principale.id)
        info = {
            'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
            'form': form,
            'projet': indicateur.activite_principale.volet.domaine.projet,
            'indicateur': indicateur,
        }
        return render(request, 'office/docprojet/edit_indicateur.html', context=info)
    except Exception as exc:
        return HttpResponse(exc)


def mod_modifier_indicateur(request, pk):
    try:
        mesures_securitaires(request)
        if request.method == 'POST':
            indic = models.Suivi_de_indicateurs_projet.objects.get(id=pk)
            form = formulaire.Suivi_des_indicateurs(request.POST, instance=indic)
            if form.is_valid():
                form.save()
                return redirect('apercu_du_projet', indic.activite_principale.volet.domaine.projet.id)
    except Exception as exc:
        return HttpResponse(exc)


def mod_delete_indicateur(request, pk):
    mesures_securitaires(request)
    indic = models.Suivi_de_indicateurs_projet.objects.get(id=pk)
    indic.delete()
    return redirect('apercu_du_projet', indic.activite_principale.volet.domaine.projet.id)


def edit_risque(request, pk):
    mesures_securitaires(request)
    risque = models.Gestion_des_risques_SI.objects.get(id=pk)
    form = formulaire.Gestion_des_risques(instance=risque)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'form': form,
        'projet': risque.volet.domaine.projet,
        'risque': risque,
    }
    return render(request, 'office/docprojet/edit_risque.html', context=info)


def mod_delete_risque(request,pk):
    mesures_securitaires(request)
    risque = models.Gestion_des_risques_SI.objects.get(id=pk)
    risque.delete()
    return redirect('apercu_du_projet', risque.volet.domaine.projet.id)


#BUDGET ET PLAN DE DECAISSEMENT
def budget(request, pk):
    projet = models.Doc_Projet.objects.get(id=pk)
    activite_principales = models.ActivitePrincipal.objects.all()
    partenaire = models.PartenaireProjet.objects.filter(projet=projet)
    n_partenaire = len(partenaire)
    total_qte = 0
    total_frequence = 0
    total_pu = 0
    total_pt = 0
    contrib_totaux = {}
    for partner in partenaire:
        contrib_totaux['{}'.format(partner.partenaire)] = 0
    for budget in models.BudgetEtPlanDecaissement.objects.all():
        if budget.activite_principale.volet.domaine.projet == projet:
            for partener, contribution in budget.partenaire:
                partner_name = f"{partener}"
                contrib_totaux[partner_name] += int(contribution)

    for budget in models.BudgetEtPlanDecaissement.objects.all():
        if budget.activite_principale.volet.domaine.projet == projet:
            total_qte = total_qte + budget.quantite
            total_frequence = total_frequence + budget.frequence
            total_pu = total_pu + budget.prixUnit
            total_pt = total_pt + budget.prixTot


    liste_budget = models.BudgetEtPlanDecaissement.objects.all()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'projet':projet,
        'activite_principale': activite_principales,
        'liste_budget': liste_budget,
        'total_qte': total_qte,
        'total_frequence': total_frequence,
        'total_pu': total_pu,
        'total_pt': total_pt,
        'l_partenaire': partenaire,
        'n_part': n_partenaire,
        'contrib_totaux': contrib_totaux
    }
    return render(request,'office/docprojet/budget_plan_decaissement/budget.html',context=info)


def mod_ajouter_budgetPlan(request, pk):
    if request.method == 'POST':
        list_partenaire = []
        for key, value in request.POST.items():
            if key.startswith("partenaire_"):
                list_partenaire.append((key[11:], value))

        idactp = request.POST['activite_principale']
        actp = models.ActivitePrincipal.objects.get(id=idactp)
        budgetplan = models.BudgetEtPlanDecaissement.objects.create(
            activite_principale=actp,
            numLigne=request.POST['num_ligne_budgetaire'],
            designation=request.POST['designation'],
            specification=request.POST['specification'],
            unite=request.POST['unite'],
            quantite=request.POST['quantite'],
            frequence=request.POST['prix_unitaire'],
            prixUnit=request.POST['frequence'],
            partenaire=list_partenaire,
            observation=request.POST['observation'],
        )
        budgetplan.save()
        bg = models.BudgetEtPlanDecaissement.objects.filter(activite_principale=actp)
        mis_a_jour_budget(bg[0].activite_principale.volet.domaine.projet.id)
        return redirect('budget', pk)
    else:
        return redirect('index_office')


def delete_budget(request,pk):
    mesures_securitaires(request)
    budget_a_supprimer = models.BudgetEtPlanDecaissement.objects.get(id=pk)
    budget_a_supprimer.delete()
    return redirect("budget", budget_a_supprimer.activite_principale.volet.domaine.projet.id)


def edit_budget(request, pk):
    mesures_securitaires(request)
    budget = models.BudgetEtPlanDecaissement.objects.get(id=pk)
    form = formulaire.BudgetPlan(instance=budget)
    form.fields['activite_principale'].widget = widgets.HiddenInput()
    partenaire_contrib = budget.partenaire
    info = {
        'form':form,
        'budget':budget,
        'partenaire_contrib': partenaire_contrib,
    }
    return render(request,'office/docprojet/budget_plan_decaissement/edit_budget.html',context=info)


def mod_modifier_budget(request, pk):
    if request.method == 'POST':
        budget = models.BudgetEtPlanDecaissement.objects.get(id=pk)
        form = formulaire.BudgetPlanMod(data=request.POST, instance=budget)
        if form.is_valid():
            myform = form.save(commit=False)
            list_partenaire = []
            for key, value in request.POST.items():
                if key.startswith("partenaire_"):
                    list_partenaire.append((key[11:], value))
            myform.partenaire = list_partenaire
            myform.save()
        else:
            return HttpResponse("Erreur form {}".format(form.errors))
        return redirect('budget', budget.activite_principale.volet.domaine.projet.id)
    else:
        return redirect('index_office')


def mis_a_jour_budget(pk):
    """Cette fonction permet de mettre a jours le prix total
        de chaque enregistrement du bugdet
    """
    projet = models.Doc_Projet.objects.get(id=pk)
    budget_a_modifier = models.BudgetEtPlanDecaissement.objects.all()
    if len(budget_a_modifier):
        for bg in budget_a_modifier:
            if bg.activite_principale.volet.domaine.projet == projet:
                bg.prixTot = bg.prixUnit * bg.quantite * bg.frequence
                bg.save()

# def mis_a_plan_detaille(projet:int):
#     """Cette fonction permet de mettre à jour la duree dans l'enregistrement
#     du plan detaille
#     """
#     plan = models.Planification_Detaille_Execution.objects.filter(projet=projet)
#     # etat_de_besoin = models.EtatdeBesoin_LigneBudgetaire.objects.filter(projet=projet)
#     # budget = models.BudgetEtPlanDecaissement.objects.filter(projet=projet)
#     for p in plan:
#         delta = abs(p.date_fin - p.date_debut)
#         # delta = format(delta, locale='fr_FR')
#         string = "{}".format(delta)
#         string = string.replace(', 0:00:00', '')
#         if 'day' in string:
#             string = string.replace('day', 'jour')
#         elif 'days' in string:
#             string = string.replace('days', 'jours')
#         p.duree = string
#         p.save()
    # for etat in etat_de_besoin:
    #     for bg in budget:
    #         if etat.sous_activite == bg.activite:
    #             etat.ligne_budgetaire = bg.numLigne
    #             etat.save()

def mis_a_jour_duree_projet(projet:int):
    projet = models.Doc_Projet.objects.get(id=projet)
    delta = projet.date_fin - projet.date_debut
    delta = delta.days
    if delta < 31:
        duree = "{} jour(s)".format(delta)
    else:
        delta = divmod(delta,31)
        duree = "{} mois et {} jour(s)".format(delta[0],delta[1])
    projet.duree = duree
    projet.save()

def planification_des_activites(request,pk):
    mesures_securitaires(request)
    indicateurs = models.Suivi_de_indicateurs_projet.objects.all()
    projet = models.Doc_Projet.objects.get(id=pk)
    domaine = models.DomaineProjet.objects.filter(projet=projet)
    voletprojet = models.VoletProjet.objects.all()
    activitePrincipal = models.ActivitePrincipal.objects.all()
    formulaire_sa = formulaire.PlanDetaille_Sa()
    sa = models.Sous_activite_Ap.objects.all()
    personnel_execution = models.Personnel_de_mise_en_oeuvre.objects.all()
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'projet': projet,
        'domaineprojet': domaine,
        'voletprojet': voletprojet,
        'activiteprincipal': activitePrincipal,
        'indicateurs': indicateurs,
        'formulaire_sa': formulaire_sa,
        'sous_activite': sa,
        'personnel_execution': personnel_execution,
    }
    return render(request,'office/docprojet/planification_des_activites/planification_des_activites.html', context=info)

def ajouter_personnel(request):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Personnel_Execution(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({"message": "ajouté"})
        else:
            return JsonResponse({"message": "le formulaire est invalide"})
    else:
        return JsonResponse({'message': 'Erreur : méthode non autorisée.'}, status=405)



def module_ajouter_une_planification(request):
    mesures_securitaires(request)
    try:
        if request.method == 'POST':
            form = formulaire.PlanDetaille_Sa(request.POST)
            if form.is_valid():
                print(2)
                indicateur = models.Suivi_de_indicateurs_projet.objects.get(id=request.POST['indicateur_id'])
                print(3)
                print(indicateur)
                myform = form.save(commit=False)
                myform.indicateur = indicateur
                myform.save()

                return JsonResponse({"message": "Le formulaire a été soumis"})
            else:
                return JsonResponse({"message": "le formulaire est invalide"})
        else:
            return JsonResponse({'message': 'Erreur : méthode non autorisée.'}, status=405)
    except Exception as exc:
        return JsonResponse({'message': exc})


def edit_plan_detail(request, pk):
    mesures_securitaires(request)
    plan = models.Sous_activite_Ap.objects.get(id=pk)
    info = {
        'plan': plan,
    }
    return render(request,"office/docprojet/planification_des_activites/edit_plan_detail.html", context=info)


def modifier_sa_detail(request, pk):
    mesures_securitaires(request)
    sa_detail = models.Sous_activite_Ap.objects.get(id=pk)
    indic = sa_detail.indicateur.id
    form = formulaire.Modif_sa_detail(key=indic, instance=sa_detail)
    info = {
        'form': form,
        'sa_detail': sa_detail,
    }
    return render(request, "office/docprojet/planification_des_activites/modif_sa_detail.html", context=info)


def mod_modifier_sa_detail(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        sa = models.Sous_activite_Ap.objects.get(id=pk)
        form = formulaire.Modif_sa_detail(data=request.POST, instance=sa)
        if form.is_valid():
            form.save()
            return redirect('planification_detaillee', sa.indicateur.activite_principale.volet.domaine.projet.id)
        else:
            return http.HttpResponse("Erreur form  : {}".format(form.errors))
    else:
        return redirect('index_office')


def mod_modifier_plan(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        pass
        # plan = models.Planification_Detaille_Execution.objects.get(id=pk)
        # form = formulaire.Planification_Execution(data=request.POST, instance=plan)
        # if form.is_valid():
        #     form.save()
        #     return redirect('planification_detaillee', request.POST['projet'])
        # else:
        #     return http.HttpResponse("Erreur form  : {}".format(form.errors))
    else:
        return redirect('index_office')


def delete_plan_detail(request, pk, idprojet):
    mesures_securitaires(request)
    plan = models.Planification_Detaille_Execution.objects.get(id=pk)
    indic = plan.indicateur
    sousactivite = models.Sous_activite_Ap.objects.filter(indicateur=indic)
    for sa in sousactivite:
        sa.delete()
    plan.delete()
    return redirect("planification_detaillee", idprojet)

def delete_sous_activite_detail(request, pk, idprojet):
    mesures_securitaires(request)
    sa = models.Sous_activite_Ap.objects.get(id=pk)
    sa.delete()
    return redirect("planification_detaillee", idprojet)


def mod_ajouter_bulletin_info(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Bulletin_Info(request.POST)
        if form.is_valid():
            myForm = form.save(commit=False)
            myForm.projet = models.Doc_Projet.objects.get(id=pk)
            myForm.save()
            return redirect("bulletinInfo",pk)
    else:
        return redirect("index_office")


def mod_ajouter_photo_realisation(request,pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Photo_Realisation_Sitrep(data=request.POST,files=request.FILES)
        if form.is_valid():
            projet = models.Doc_Projet.objects.get(id=pk)
            if len(models.BulletinInfo.objects.filter(projet=projet)) > 0:
                form.save()
                return redirect("bulletinInfo",pk)
            else:
                return http.HttpResponse("Il faut inserer d'abord un bulletin")
        else:
            return HttpResponse("Erreur form {}: {} ".format(form.errors,request.POST))
    else:
        return redirect("index_office")

def exporter_doc_projet(request, pk):
    try:
        projet = models.Doc_Projet.objects.get(id=pk)
        doc = DocxTemplate(os.getcwd() + "/media/outils/doc_projet.docx")
        context = {'titre': projet.titre, 'code': projet.code_project,
                   'finance_par': projet.finance_par, 'date_debut': projet.date_debut, 'date_fin': projet.date_fin,
                   'duree': projet.duree, 'plan_action_lac_dvlpmt': projet.plan_action, 'programme_sollicite': projet.programme_solicite,
                   'context_actualise': projet.contextActualiseProgramme, 'description_impact': projet.descImpactAttenduProg}

        doc.render(context)
        doc.save("doc_projet_tmpl.docx")

        pythoncom.CoInitialize()
        word = win32.gencache.EnsureDispatch('Word.Application')
        document = word.Documents.Open(os.getcwd() + '/doc_projet_tmpl.docx')

        paysprojet = models.Pays_Projet.objects.filter(projet=projet)
        paysprojetcount = len(paysprojet)
        table = document.Tables(3)
        if paysprojet:
            for i in range(1, 6):
                cell = table.Cell(2, i)
                cell.Split(paysprojetcount, 1)

            paysindex = 2
            for pays in paysprojet:
                mycell = table.Cell(paysindex, 1)
                mycell.Range.Text = pays.pays

                provinceprojet = models.Province_Projet.objects.filter(pays=pays)
                provincecount = len(provinceprojet)

                if provinceprojet:
                    for i in range(2, 6):
                        cell2 = table.Cell(mycell.RowIndex, i)
                        cell2.Split(provincecount, 1)

                    provincerowindex = mycell.RowIndex
                    for province in provinceprojet:
                        mycell = table.Cell(provincerowindex, 2)
                        mycell.Range.Text = province.province

                        territoireprojet = models.Ville_Territoire_Projet.objects.filter(province=province)
                        territoirecount = len(territoireprojet)

                        if territoireprojet:
                            for i in range(3, 6):
                                cell3 = table.Cell(mycell.RowIndex, i)
                                cell3.Split(territoirecount, 1)

                            territoireindex = mycell.RowIndex
                            for territoire in territoireprojet:
                                mycell = table.Cell(territoireindex, 3)
                                mycell.Range.Text = territoire.ville_territoire

                                axeprojet = models.Axe_Projet.objects.filter(ville_territoire=territoire)
                                axecount = len(axeprojet)

                                if axeprojet:
                                    for i in range(4, 6):
                                        cell4 = table.Cell(mycell.RowIndex, i)
                                        cell4.Split(axecount, 1)

                                    axeindex = mycell.RowIndex
                                    for axe in axeprojet:
                                        mycell = table.Cell(axeindex, 4)
                                        mycell.Range.Text = axe.axe
                                        mycell = table.Cell(axeindex, 5)
                                        mycell.Range.Text = axe.nombre_beneficiares

                                        axeindex += 1
                                    territoireindex += axecount
                                    provincerowindex = territoireindex
                                else:
                                    territoireindex += 1
                                    provincerowindex = territoirecount
                            paysindex = provincerowindex
                        else:
                            provincerowindex += 1
                            paysindex = provincecount
                else:
                    paysindex += 1


        coordnat = models.Coordination_Nationale_Projet.objects.filter(projet=projet)
        coordnatcount = len(coordnat)
        table = document.Tables(4)
        if coordnat:
            for i in range(1, 4):
                cell = table.Cell(2, i)
                cell.Split(coordnatcount, 1)
            coordnatindex = 2
            for cn in coordnat:
                mycell = table.Cell(coordnatindex, 1)
                mycell.Range.Text = cn.coordination_national

                coordprov = models.Coordination_Provinciale_Projet.objects.filter(cordination_nationale=cn)
                coordprovcount = len(coordprov)
                if coordprov:
                    for i in range(2, 4):
                        cell2 = table.Cell(mycell.RowIndex, i)
                        cell2.Split(coordprovcount, 1)
                    coordprovrowindex = mycell.RowIndex
                    for cp in coordprov:
                        mycell = table.Cell(coordprovrowindex, 2)
                        mycell.Range.Text = cp.coordination_provinciale
                        antenneprojet = models.Antenne_Projet.objects.filter(coordination_provinciale=cp)
                        antennecount = len(antenneprojet)
                        if antenneprojet:
                            cell3 = table.Cell(mycell.RowIndex, 3)
                            cell3.Split(antennecount, 1)
                            antenneindex = mycell.RowIndex
                            for antenne in antenneprojet:
                                mycell = table.Cell(antenneindex, 3)
                                mycell.Range.Text = antenne.antenne_projet

                                antenneindex += 1
                            coordprovrowindex += antennecount
                            coordnatindex = coordprovrowindex
                        else:
                            coordprovrowindex += 1
                            coordnatindex = coordprovrowindex
                else:
                    coordnatindex += 1

        domaineprojet = models.DomaineProjet.objects.filter(projet=projet)
        domainecount = len(domaineprojet)
        if domaineprojet:
            indextable = 6
            copytable = document.Tables(5)
            for domaine in domaineprojet:
                table = document.Tables(indextable)
                mycell = table.Cell(1, 1)
                mycell.Range.Text = "Domaine des Résultats : " + domaine.domaine.domaine
                mycell = table.Cell(2, 1)
                mycell.Range.Text = "Objectif LACasbl de développement dans ce domaine : " + domaine.objectif_lac_de_dvpt_dans_ce_domaine
                voletdomaine = models.VoletProjet.objects.filter(domaine=domaine)
                if voletdomaine:
                    voletindex = 3
                    for volet in voletdomaine:
                        copytable.Range.Select()
                        # Copier le tableau dans le presse-papiers
                        word.Selection.Copy()
                        # Déplacer le curseur à la fin du tableau existant
                        document.Range(table.Range.End, table.Range.End).Select()
                        # Coller le tableau à partir de la dernière ligne
                        word.Selection.Paste()
                        mycell = table.Cell(voletindex, 1)
                        mycell.Range.Text = "VOLET : " + volet.volet.volet
                        mycell = table.Cell(voletindex + 1, 1)
                        mycell.Range.Text = "OBJECTIF LACAsbl DE DEVELOPPEMENT AUQUEL LE PROGRAMME CONTRIBUE DANS CE VOLET : " + volet.objectif_LAC
                        mycell = table.Cell(voletindex + 2, 1)
                        mycell.Range.Text = "OBJECTIF SPECIFIQUE SMART POURSUIVI PAR CE PROGRAMME DANS CE VOLET : " + volet.objectif_SMART
                        mycell = table.Cell(voletindex + 3, 1)
                        mycell.Range.Text = "VUE D'ENSEMBLE OU SOMMAIRE DE LA STRATEGIE ET TACTIQUE POUR ATTEINDRE L'OBJECTIF DU PROGRAMME DANS CE VOLET : " + volet.vue_d_ensemble
                        mycell = table.Cell(voletindex + 4, 1)
                        mycell.Range.Text = "DESCRIPTION DES EFFETS ESPERES : " + volet.description_des_effets
                        mycell = table.Cell(voletindex + 5, 1)
                        mycell.Range.Text = "NOMBRE TOTAL D'HISTOIRE DE SUCCES A DOCUMENTER : " + str(volet.nbre_succes)
                        mycell = table.Cell(voletindex + 6, 1)
                        mycell.Range.Text = "ENUMERER LE PERSONNEL CLE ET LE NOMBRE DES JOURS DE TRAVAIL : " + volet.enumererpersonnel

                        actpindex = voletindex + 9
                        num = 1
                        actp = models.ActivitePrincipal.objects.filter(volet=volet)
                        countactp = len(actp)
                        if actp:
                            for i in range(1, 21):
                                cellactp = table.Cell(actpindex, i)
                                cellactp.Split(countactp, 1)
                            for ap in actp:
                                mycell = table.Cell(actpindex, 1)
                                mycell.Range.Text = num
                                mycell = table.Cell(actpindex, 2)
                                mycell.Range.Text = ap.activitePrincipale
                                mycell = table.Cell(actpindex, 3)
                                mycell.Range.Text = ap.responsable_execution
                                indicateurs = models.Suivi_de_indicateurs_projet.objects.filter(activite_principale=ap)
                                indicateurcount = len(indicateurs)
                                num+=1
                                if indicateurs:
                                    for i in range(4, 21):
                                        cell1 = table.Cell(mycell.RowIndex, i)
                                        cell1.Split(indicateurcount, 1)
                                    indicindex = mycell.RowIndex
                                    for indicateur in indicateurs:
                                        mycell = table.Cell(indicindex, 4)
                                        mycell.Range.Text = indicateur.indicateur_lac
                                        mycell = table.Cell(indicindex, 5)
                                        mycell.Range.Text = indicateur.base_line_gab
                                        mycell = table.Cell(indicindex, 6)
                                        mycell.Range.Text = indicateur.cible_NA
                                        mycell = table.Cell(indicindex, 7)
                                        mycell.Range.Text = indicateur.source_outils
                                        mycell = table.Cell(indicindex, 8)
                                        mycell.Range.Text = indicateur.frequence_de_collecte
                                        if indicateur.periode_execution:
                                            for mois in indicateur.periode_execution:
                                                if mois == "janvier":
                                                    mycell1 = table.Cell(indicindex, 9)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "février":
                                                    mycell1 = table.Cell(indicindex, 10)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "mars":
                                                    mycell1 = table.Cell(indicindex, 11)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "avril":
                                                    mycell1 = table.Cell(indicindex, 12)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "mai":
                                                    mycell1 = table.Cell(indicindex, 13)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "juin":
                                                    mycell1 = table.Cell(indicindex, 14)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "juillet":
                                                    mycell1 = table.Cell(indicindex, 15)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "aout":
                                                    mycell1 = table.Cell(indicindex, 16)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "septembre":
                                                    mycell1 = table.Cell(indicindex, 17)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "octobre":
                                                    mycell1 = table.Cell(indicindex, 18)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "novembre":
                                                    mycell1 = table.Cell(indicindex, 19)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                                elif mois == "décembre":
                                                    mycell1 = table.Cell(indicindex, 20)
                                                    mycell1.Shading.BackgroundPatternColor = win32.constants.wdColorTeal
                                        indicindex += 1
                                    actpindex += indicateurcount
                                else:
                                    actpindex += 1
                        indexgestion = actpindex + 2
                        risques = models.Gestion_des_risques_SI.objects.filter(volet=volet)
                        countrisque = len(risques)
                        numrisque = 1
                        if risques:
                            for risque in risques:
                                mycell = table.Cell(indexgestion, 1)
                                mycell.Range.Text = numrisque
                                mycell = table.Cell(indexgestion, 2)
                                mycell.Range.Text = risque.description
                                mycell = table.Cell(indexgestion, 3)
                                mycell.Range.Text = risque.mesure
                                mycell = table.Cell(indexgestion, 4)
                                mycell.Range.Text = risque.responsable
                                indexgestion += 1
                                numrisque += 1
                                if numrisque <= countrisque:
                                    table.Rows.Add()
                        voletindex = indexgestion
                indextable += 1
            table = document.Tables(5)
            table.Delete()
            for i in range(10-domainecount):
                table = document.Tables(5+domainecount)
                table.Delete()

        document.Save()
        document.Close()
        word.Quit()

    except Exception as exc:
        document.Close()
        word.Quit()
        print(exc)

    return HttpResponse("Exporté")


def exporter_doc_budget(request, pk):
    try:
        projet = models.Doc_Projet.objects.get(id=pk)
        partenaires = models.PartenaireProjet.objects.filter(projet=projet)
        n_partenaires = len(partenaires)
        activites = models.ActivitePrincipal.objects.all()
        workbook = openpyxl.load_workbook(os.getcwd() + '/media/outils/budget_global.xlsx')
        worksheet = workbook.active

        list_alpha = ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB',
                      'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR',
                      'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

        cell = worksheet['E1']
        cell.value = projet.titre
        cell = worksheet['E2']
        cell.value = projet.code_project
        celluleindex = 7
        thcontrib = 6
        worksheet.merge_cells('J5:{}5'.format(list_alpha[n_partenaires - 1]))
        worksheet.merge_cells('{}5:{}6'.format(list_alpha[n_partenaires], list_alpha[n_partenaires]))
        cell = worksheet['{}5'.format(list_alpha[n_partenaires])]
        cell.value = "OBSERVATIONS"
        if activites:
            for act in activites:
                if act.volet.domaine.projet == projet:
                    cell = worksheet['A{}'.format(celluleindex)]
                    cell.value = act.activitePrincipale
                    budget = models.BudgetEtPlanDecaissement.objects.filter(activite_principale=act)
                    countbudget = len(budget)
                    if budget:
                        if len(budget) > 1:
                            worksheet.merge_cells('A{}:A{}'.format(celluleindex, celluleindex + countbudget - 1))
                        budgetcellcount = celluleindex
                        for bdgt in budget:
                            cell = worksheet['B{}'.format(budgetcellcount)]
                            cell.value = bdgt.numLigne
                            cell = worksheet['C{}'.format(budgetcellcount)]
                            cell.value = bdgt.designation
                            cell = worksheet['D{}'.format(budgetcellcount)]
                            cell.value = bdgt.specification
                            cell = worksheet['E{}'.format(budgetcellcount)]
                            cell.value = bdgt.unite
                            cell = worksheet['F{}'.format(budgetcellcount)]
                            cell.value = float(bdgt.quantite)
                            cell = worksheet['G{}'.format(budgetcellcount)]
                            cell.value = float(bdgt.frequence)
                            cell = worksheet['H{}'.format(budgetcellcount)]
                            cell.value = float(bdgt.prixUnit)
                            i = 0
                            for part_contrib in bdgt.partenaire:
                                cell = worksheet['{}6'.format(list_alpha[i], thcontrib)]
                                cell.value = part_contrib[0]
                                cell = worksheet['{}{}'.format(list_alpha[i], budgetcellcount)]
                                cell.value = float(part_contrib[1])
                                i += 1
                            cell = worksheet['{}{}'.format(list_alpha[i], budgetcellcount)]
                            cell.value = bdgt.observation
                            budgetcellcount += 1
                        celluleindex = budgetcellcount
                    else:
                        celluleindex += 1

            workbook.save('doc_budget_tmpl.xlsx')
            messages.success(request, 'Exporté')
        return redirect('budget', pk)
    except Exception as exc:
        return HttpResponse(exc)


def exporter_doc_plan_detaille(request, pk):
    try:
        projet = models.Doc_Projet.objects.get(id=pk)
        volets = models.VoletProjet.objects.all()
        sous_activite = models.Sous_activite_Ap.objects.all()
        list_volet = []
        for volet in volets:
            if volet.domaine.projet == projet:
                list_volet.append(volet)
        if list_volet:
            workbook = openpyxl.load_workbook(os.getcwd() + '/media/outils/planification_detaillee.xlsx')
            worksheet = workbook.active
            celluleindex = 8
            for volet in list_volet:
                cell = worksheet['A{}'.format(celluleindex)]
                cell.value = volet.objectif_SMART
                actp = models.ActivitePrincipal.objects.filter(volet=volet)
                countactp = len(actp)
                list_sa = []
                for sa in sous_activite:
                    if sa.indicateur.activite_principale.volet == volet:
                        list_sa.append(sa)
                counttot = countactp + len(list_sa)
                if counttot > 1:
                    worksheet.merge_cells('A{}:A{}'.format(celluleindex, celluleindex + counttot - 2))
                if actp:
                    actpcellindex = celluleindex
                    for ap in actp:
                        cell = worksheet['B{}'.format(actpcellindex)]
                        cell.value = ap.activitePrincipale
                        list = []
                        for sa in sous_activite:
                            if sa.indicateur.activite_principale == ap:
                                list.append(sa)
                        countsa = len(list)
                        sacellindex = actpcellindex
                        if len(list) > 1:
                            worksheet.merge_cells('B{}:B{}'.format(actpcellindex, actpcellindex + countsa - 1))
                        if list:
                            for sa in list:
                                cell = worksheet['C{}'.format(sacellindex)]
                                cell.value = sa.sousactivite
                                cell = worksheet['D{}'.format(sacellindex)]
                                cell.value = sa.descapproche
                                cell = worksheet['E{}'.format(sacellindex)]
                                cell.value = sa.strategie_genre
                                cell = worksheet['F{}'.format(sacellindex)]
                                cell.value = sa.inclusion_pers_besoin
                                cell = worksheet['G{}'.format(sacellindex)]
                                cell.value = sa.contribution
                                cell = worksheet['H{}'.format(sacellindex)]
                                cell.value = sa.besoinenappui
                                cell = worksheet['I{}'.format(sacellindex)]
                                cell.value = sa.outildesupervision
                                cell = worksheet['J{}'.format(sacellindex)]
                                cell.value = sa.partenairetechnique
                                cell = worksheet['K{}'.format(sacellindex)]
                                cell.value = sa.partenairetatique
                                cell = worksheet['L{}'.format(sacellindex)]
                                cell.value = sa.lieu_execution
                                cell = worksheet['M{}'.format(sacellindex)]
                                cell.value = sa.date_debut
                                cell = worksheet['N{}'.format(sacellindex)]
                                cell.value = sa.date_fin
                                cell = worksheet['O{}'.format(sacellindex)]
                                cell.value = sa.date_rapportage
                                sacellindex += 1
                            actpcellindex = sacellindex
                        else:
                            actpcellindex += 1
                    celluleindex = actpcellindex
                else:
                    celluleindex += 1
            workbook.save('planification_detaillee_tmpl.xlsx')
        else:
            return HttpResponse("Veuillez d'abord terminer la configuration du projet")

        return HttpResponse("Exporté")
    except Exception as exc:
        return HttpResponse(exc)


def les_volets(request):
    try:
        domaine = models.Domaine.objects.get(domaine=request.GET.get('domaine'))
        volets = models.Volet.objects.filter(domaine=domaine).values()

        return JsonResponse(list(volets), safe=False)
    except Exception as exc:
        return HttpResponse(exc)


def les_indicateurs(request):
    try:
        activiteprincipale = models.ActivitePrincipal.objects.get(id=request.GET.get('ap'))
        indicateurs = models.Suivi_de_indicateurs_projet.objects.filter(activite_principale=activiteprincipale).values('id', 'indicateur_lac')
        return JsonResponse(list(indicateurs), safe=False)
    except Exception as exc:
        return HttpResponse(exc)


def les_lignes_budgetaires(request):
    try:
        indicateur = models.Suivi_de_indicateurs_projet.objects.get(id=request.GET.get('indicateur'))
        ap = indicateur.activite_principale
        ligne = models.BudgetEtPlanDecaissement.objects.filter(activite_principale=ap).values("numLigne")
        return JsonResponse(list(ligne), safe=False)
    except Exception as exc:
        return HttpResponse(exc)


def mod_ajouter_domaine_configuration(request, pk):
    try:
        projet = models.Doc_Projet.objects.get(id=pk)
        domaine = models.Domaine.objects.get(id=request.POST['domaine'])
        scan = models.DomaineProjet.objects.filter(projet=projet, domaine=domaine)
        if scan:
            return HttpResponse("Ce domaine a déjà été configuré pour ce projet")
        else:
            domaineprojet = models.DomaineProjet.objects.create(projet=projet, domaine=domaine, objectif_lac_de_dvpt_dans_ce_domaine=
                                                    request.POST['objectiflacdomaine'])
            domaineprojet.save()
            mydomaineprojet = models.DomaineProjet.objects.get(projet=projet, domaine=domaine)
            for key, value in request.POST.items():
                if key.startswith('volet_'):
                    monvolet = models.Volet.objects.get(volet=str(value))
                    for keyobjdvlpmtlac, valueobjdvlpmtlac in request.POST.items():
                        if keyobjdvlpmtlac == 'objectiflac_{}'.format(key):
                            volet_objectif_LAC = valueobjdvlpmtlac
                    for keyobjspecsmart, valueobjspecsmart in request.POST.items():
                        if keyobjspecsmart == 'objectifspecsmart_{}'.format(key):
                            volet_objectif_SMART = valueobjspecsmart
                    for keyvuestrategie, valuevuestrategie in request.POST.items():
                        if keyvuestrategie == 'vuestrategietactique_{}'.format(key):
                            volet_vue_d_ensemble = valuevuestrategie
                    for keydescefspr, valuedescefspr in request.POST.items():
                        if keydescefspr == 'descripeffetspr_{}'.format(key):
                            volet_description_des_effets = valuedescefspr
                    for keynbresucces, valuenbresucces in request.POST.items():
                        if keynbresucces == 'nombreTotalHist_{}'.format(key):
                            volet_nbre_succes = valuenbresucces
                    for keyenumpersonnel, valueenumpersonnel in request.POST.items():
                        if keyenumpersonnel == 'enumererPersonnelCle_{}'.format(key):
                            volet_enumererpersonnel = valueenumpersonnel


                    myvoletprojet = models.VoletProjet.objects.create(domaine=mydomaineprojet, volet=monvolet,
                                    objectif_LAC=volet_objectif_LAC, objectif_SMART=volet_objectif_SMART,
                                    vue_d_ensemble=volet_vue_d_ensemble, description_des_effets=volet_description_des_effets,
                                    nbre_succes=volet_nbre_succes, enumererpersonnel=volet_enumererpersonnel)
                    myvoletprojet.save()
                    voletprojet = models.VoletProjet.objects.get(domaine=mydomaineprojet, volet=monvolet)
                    volet_activitePrincipale = ""
                    volet_responsable_execution = ""
                    for keyap, valueap in request.POST.items():
                        if keyap.startswith('activite_'):
                            if keyap.endswith('{}'.format(key)):
                                volet_activitePrincipale = str(valueap)
                                for keyresponsablexec, valueresponsablexec in request.POST.items():
                                    if keyresponsablexec == 'responsable_{}'.format(keyap):
                                        volet_responsable_execution = str(valueresponsablexec)
                                myactiviteprincipale = models.ActivitePrincipal.objects.create(volet=voletprojet,
                                            activitePrincipale=volet_activitePrincipale, responsable_execution=volet_responsable_execution)
                                myactiviteprincipale.save()
                                activiteprincipale = models.ActivitePrincipal.objects.get(volet=voletprojet, activitePrincipale=volet_activitePrincipale)

                                activite_indicateur_lac = ""
                                activite_base_line_gab = ""
                                activite_cible_NA = ""
                                activite_source_outils = ""
                                activite_frequence_de_collecte = ""
                                activite_periode_execution = ""
                                for keyindicateur, valueindic in request.POST.items():
                                    if keyindicateur.startswith('indicateur_'):
                                        if keyindicateur.endswith('{}'.format(keyap)):
                                            activite_indicateur_lac = str(valueindic)
                                            for keybaseline, valuebaseline in request.POST.items():
                                                if keybaseline == 'baseline_{}'.format(keyindicateur):
                                                    activite_base_line_gab = str(valuebaseline)
                                            for keycible, valuecible in request.POST.items():
                                                if keycible == 'cible_{}'.format(keyindicateur):
                                                    activite_cible_NA = str(valuecible)
                                            for keysrcdonnees, valuesrcdonnees in request.POST.items():
                                                if keysrcdonnees == 'sourcedesdonnees_{}'.format(keyindicateur):
                                                    activite_source_outils = str(valuesrcdonnees)
                                            for keyfrequence, valuefrequence in request.POST.items():
                                                if keyfrequence == 'frequence_{}'.format(keyindicateur):
                                                    activite_frequence_de_collecte = str(valuefrequence)
                                            for keyperiode, valueperiode in request.POST.items():
                                                if keyperiode == 'periode_{}'.format(keyindicateur):
                                                    activite_periode_execution = request.POST.getlist(keyperiode)

                                            myindicateur = models.Suivi_de_indicateurs_projet.objects.create(activite_principale=activiteprincipale,
                                                    indicateur_lac=activite_indicateur_lac, base_line_gab=activite_base_line_gab, cible_NA=activite_cible_NA,
                                                    source_outils=activite_source_outils, frequence_de_collecte=activite_frequence_de_collecte,
                                                    periode_execution=activite_periode_execution)
                                            myindicateur.save()
                    gestionrisque_description = ""
                    gestionrisque_mesure = ""
                    gestionrisque_responsable = ""
                    for keygestion, valuegestion in request.POST.items():
                        if keygestion.startswith('description_'):
                            if keygestion.endswith('{}'.format(key)):
                                gestionrisque_description = str(valuegestion)
                                for keymesure, valuemesure in request.POST.items():
                                    if keymesure == 'mesuremitigation_{}'.format(keygestion):
                                        gestionrisque_mesure = str(valuemesure)
                                for keyresponsableprev, valueresponsableprev in request.POST.items():
                                    if keyresponsableprev == 'responsableprevention_{}'.format(keygestion):
                                        gestionrisque_responsable = str(valueresponsableprev)

                                gestionrisque = models.Gestion_des_risques_SI.objects.create(volet=voletprojet, description=gestionrisque_description,
                                                                                     mesure=gestionrisque_mesure, responsable=gestionrisque_responsable)
                                gestionrisque.save()
            return redirect('configuration_domaine_doc_projet', pk)

    except Exception as exc:
        return HttpResponse(exc)


def mapping_index(request):
    mesures_securitaires(request)
    info = {
        'user':models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'titre': 'Mapping - ',
        'projets': models.Doc_Projet.objects.all(),
    }
    return render(request,'office/docprojet/mapping/mapping_index.html',context=info)


def mapping_afficher(request, pk):
    mesures_securitaires(request)
    info = {
        'user':models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'titre': 'Mapping - ',
        'mapping': models.Mapping_Projet.objects.all(),
        'projet': models.Doc_Projet.objects.get(id=pk),
    }
    return render(request,'office/docprojet/mapping/mapping_affichage.html',context=info)


def nouveau_mapping(request, pk):
    form = formulaire.Mapping()
    form.initialisation(pk=pk)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'titre': 'Mapping - ',
        'projet': models.Doc_Projet.objects.get(id=pk),
        'form':form
    }
    return render(request,'office/docprojet/mapping/nouveau_mapping.html',context=info)


def mod_ajouter_mapping(request,pk):
    mesures_securitaires(request)
    projet = models.Doc_Projet.objects.get(id=pk)
    if request.method == 'POST':
        form = formulaire.Mapping(request.POST)
        if form.is_valid():
            myForm = form.save(commit=False)
            myForm.projet = projet
            myForm.save()
            messages.success(request,"Une ligne a été ajouté dans le document mapping du projet")
            return redirect("afficher le mapping du projet",projet.id)
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")


def modifier_lign_mapping(request,pk):
    mesures_securitaires(request)
    form = formulaire.Mapping(instance=models.Mapping_Projet.objects.get(id=pk))
    map = models.Mapping_Projet.objects.get(id=pk)
    form.initialisation(pk=models.Mapping_Projet.objects.get(id=pk).projet.id)
    info = {
        'user': models.Utilisateur.objects.get(nom_utilisateur=request.session['utilisateur']),
        'titre': 'Modifier mapping - ',
        'projet': map.projet,
        'map':map,
        'form': form
    }
    return render(request, 'office/docprojet/mapping/modifier_mapping.html', context=info)


def mod_modifier_mapping(request, pk):
    mesures_securitaires(request)
    if request.method == 'POST':
        form = formulaire.Mapping(request.POST, instance=models.Mapping_Projet.objects.get(id=pk))
        map = models.Mapping_Projet.objects.get(id=pk)
        if form.is_valid():
            form.save()
            return redirect("afficher le mapping du projet",  map.projet.id)
        else:
            return HttpResponse(form.errors)
    else:
        return redirect("index_office")


def mod_effacer_mapping(request,pk):
    mesures_securitaires(request)
    map = models.Mapping_Projet.objects.get(id=pk)
    map.delete()
    return redirect("afficher le mapping du projet",map.projet.id)



